import tkinter as tk
import numpy as np
import time
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Color, colors
from rms import RMS
from edf import EDF

# Two entries, RMS and EDF
# Input is a task set
# Task set is run through each function
# output of each is generated via GUI using Tkinter


# schedule task based on lowest priority
# schedule up to LCM of the deadlines
# Phase 1: print array of tasks in order
# Phase final: prints excel style on GUI


if __name__ == '__main__':

    while True:
        # TKinter setup for GUI of scheduling
        global e
        preview = 0
        # action = input("Please input at least 2 task sets in format Ci Pi Dl, Ci Pi Dl ")
        # choice = input("Please input if you would like rms, edf, or both by typing 'rms' 'edf' or 'both' ")
        def scheduling_setup(string):
            replace = string.replace(",", "")
            task = replace.split(" ")
            task_convert = np.array(task)
            tasks = task_convert.astype(int)
            n = int(len(task) / 3)
            m = 3
            task_sets = tasks.reshape(n, m)

            if len(tasks) < 6:
                print("Invalid task set entered")
                label = tk.Label(root, text="The task set can not be scheduled")
                label.pack()
            else:
                return task_sets


        def rms_preview():
            string_input = e.get()
            global preview
            preview = 1
            rms_scheduler(string_input)


        def rms_scheduler(string):
            global preview
            task_sets = scheduling_setup(string)
            rms = RMS(task_sets)
            least_common_multiple = rms.lcm_rms(task_sets)
            priority_order = rms.priority_order(task_sets)
            flag = rms.exact_analysis()
            if flag is False:
                label = tk.Label(root, text="The task set can not be scheduled for RMS")
                label.pack()
            else:
                rms_schedule = rms.rms_schedule(priority_order, least_common_multiple, task_sets)
                element = ''
                for i in range(len(rms_schedule)):
                    element = element + rms_schedule[i] + ' '
                if preview == 1:
                    label = tk.Label(root, text="This is RMS")
                    label2 = tk.Label(root, text=element)
                    label.pack()
                    label2.pack()
                    preview = 0
                return rms_schedule

        def edf_preview():
            string_input = e.get()
            global preview
            preview = 1
            edf_scheduler(string_input)

        def edf_scheduler(string):
            global preview
            task_sets = scheduling_setup(string)
            edf = EDF(task_sets)
            l_c_m = edf.lcm_edf(task_sets)
            p_ord = edf.priority_order(task_sets)[0]
            flag = edf.utilization_test(task_sets)
            if flag is False:
                label = tk.Label(root, text="The task set can not be scheduled for EDF")
                label.pack()
            else:
                edf_schedule = edf.edf_schedule(p_ord, l_c_m, task_sets)
                element = ''
                for i in range(len(edf_schedule)):
                    element = element + edf_schedule[i] + ' '
                if preview == 1:
                    label = tk.Label(root, text="This is EDF")
                    label2 = tk.Label(root, text=element)
                    label.pack()
                    label2.pack()
                    preview = 0
                return edf_schedule

        def print_scheduler(string, iteration, number_of_tasks, flag):
            wb = openpyxl.load_workbook('updatedResults.xlsx')
            ws = wb.sheetnames
            wr = ws[0]
            sheet = wb[wr]
            if flag == 0:
                arr = rms_scheduler(string)
            if flag == 1:
                arr = edf_scheduler(string)

            my_cell = sheet.cell(row=(1 + iteration), column=(1))
            my_cell.value = string

            if flag == 0:
                my_cell = sheet.cell(row=(2 + iteration), column=(1))
                my_cell.value = "RMS"
                my_cell.alignment = Alignment(horizontal='center')
            else:
                my_cell = sheet.cell(row=(2 + iteration), column=(1))
                my_cell.value = "EDF"
                my_cell.alignment = Alignment(horizontal='center')

            if arr is None:
                my_cell = sheet.cell(row=(1 + iteration), column=2)
                my_cell.value = "Unscheduable"
                my_cell.alignment = Alignment(horizontal='center')
                sheet_clear = wb.active
                sheet_clear.column_dimensions['B'].width = 20
                wb.save('updatedResults.xlsx')
                point_in_sheet = 3
                return point_in_sheet

            my_cell = sheet.cell(row=(1 + iteration), column=2)
            my_cell.value = "Scheduled"
            my_cell.alignment = Alignment(horizontal='center')

            red_fill = PatternFill(start_color='FFFF0000',
                                  end_color='FFFF0000',
                                  fill_type='solid')

            for i in range(len(arr)):
                for j in range(number_of_tasks):
                    string_check = 'T' + str(j + 1)
                    if arr[i] == ' ':
                        my_cell = sheet.cell(row=(number_of_tasks + 1 + iteration), column=(i + 3))
                        my_cell.value = i
                        my_cell.alignment = Alignment(horizontal='right')
                        break
                    elif arr[i] == string_check:
                        my_cell = sheet.cell(row=(j + 1 + iteration), column=(i + 4))
                        my_cell.value = arr[i]
                        my_cell.alignment = Alignment(horizontal='center')
                        my_cell = sheet.cell(row=(number_of_tasks + 1 + iteration), column=(i+3))
                        my_cell.value = i
                        my_cell.alignment = Alignment(horizontal='right')
                    else:
                        my_cell = sheet.cell(row=(j + 1 + iteration), column=(i + 4))
                        my_cell.fill = red_fill
            my_cell = sheet.cell(row=(number_of_tasks + 1 + iteration), column=(len(arr) + 3))
            my_cell.value = len(arr)
            my_cell.alignment = Alignment(horizontal='right')
            wb.save('updatedResults.xlsx')
            point_in_sheet = number_of_tasks + 2
            return point_in_sheet

        def file_schedule_setup(flag):
            start_time = time.time()
            string = e.get()
            wb = openpyxl.load_workbook(string)
            ws = wb.sheetnames
            wr = ws[0]
            sheet = wb[wr]
            rows = sheet.max_row
            cols = sheet.max_column
            iterator = 0

            wb_clear = openpyxl.Workbook()
            wb_clear.create_sheet('results')
            sheet_clear = wb_clear.active
            sheet_clear.column_dimensions['A'].width = 10
            wb_clear.save('updatedResults.xlsx')

            for i in range(rows):
                task_string = ""
                task_count = 0
                for j in range(cols):
                    if sheet.cell(row=i+1, column=j+1).value is None:
                        break
                    my_cell = sheet.cell(row=i + 1, column=j + 1)
                    task_string = task_string + my_cell.value
                    if sheet.cell(row=i + 1, column=j + 2).value is not None:
                        task_string = task_string + ', '
                    task_count += 1
                if flag == 0:
                    iterator += print_scheduler(task_string, iterator, task_count, 0)
                elif flag == 1:
                    iterator += print_scheduler(task_string, iterator, task_count, 1)
                else:
                    iterator += print_scheduler(task_string, iterator, task_count, 0)
                    iterator += print_scheduler(task_string, iterator, task_count, 1)
            print("--- %s seconds ---" % (time.time() - start_time))

        def file_schedule_setup_rms():
            file_schedule_setup(0)

        def file_schedule_setup_edf():
            file_schedule_setup(1)

        def file_schedule_setup_both():
            file_schedule_setup(2)

        root = tk.Tk()
        root.title("RMS/EDF GUI Scheduling Tool")
        root.geometry("500x300")

        e = tk.Entry(root)
        e.place(relx = 0.5,
                rely = 0.5,
                anchor = "center")
        e.focus_set()

        T = tk.Text(root, height=2, width=60)
        T.pack()
        T.insert(tk.END, "Please enter a task set in format \"Ci Pi Di, Ci Pi Di\"\nOr enter a .xlsx file")


        button = tk.Button(root, text="RMS\nPreview", command=rms_preview)
        button.config(width=10)
        button.config(height=3)
        button.place(relx = 0.0,
                     rely = 1.0,
                     anchor = 'sw')
        button2 = tk.Button(root, text="EDF\nPreview", command=edf_preview)
        button2.config(width=10)
        button2.config(height=3)
        button2.place(relx=1.0,
                     rely=1.0,
                     anchor='se')
        button3 = tk.Button(root, text="RMS to Excel\nfrom file", command=file_schedule_setup_rms)
        button3.config(width=10)
        button3.config(height=3)
        button3.place(relx=0.0,
                      rely=0.6,
                      anchor='sw')
        button4 = tk.Button(root, text="EDF to Excel\nfrom file", command=file_schedule_setup_edf)
        button4.config(width=10)
        button4.config(height=3)
        button4.place(relx=1.0,
                      rely=0.6,
                      anchor='se')
        button5 = tk.Button(root, text="RMS/EDF to Excel\nfrom file", command=file_schedule_setup_both)
        button5.config(width=15)
        button5.config(height=3)
        button5.place(relx=0.5,
                      rely=1.0,
                      anchor='s')

        root.mainloop()
